"""
Excel Export Utility
Exports violations and vehicle data to formatted .xlsx files.
Install: pip install openpyxl
"""

import os
import sqlite3
from datetime import datetime

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.abspath(os.path.join(_HERE, ".."))


def export_to_excel(db_path: str = None, output_path: str = None) -> str:
    """
    Export all violations to a formatted Excel file.
    Returns path to .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    db_path     = db_path     or os.path.join(_ROOT, "logs", "violations.db")
    output_path = output_path or os.path.join(
        _ROOT, "results",
        f"violations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Fetch data
    violations, vehicles = [], []
    if os.path.exists(db_path):
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            violations = [dict(r) for r in conn.execute(
                "SELECT * FROM violations ORDER BY id DESC"
            ).fetchall()]
            try:
                vehicles = [dict(r) for r in conn.execute(
                    "SELECT * FROM all_vehicles ORDER BY id DESC LIMIT 500"
                ).fetchall()]
            except Exception:
                pass

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    DARK_BLUE  = "FF050911"
    ACCENT     = "FF4DB8FF"
    RED        = "FFFF4466"
    AMBER      = "FFFFAA22"
    GREEN      = "FF22DD88"
    WHITE      = "FFFFFFFF"
    LIGHT_ROW  = "FFF0F4F8"

    header_font  = Font(name="Calibri", bold=True, color=ACCENT,  size=10)
    title_font   = Font(name="Calibri", bold=True, color=WHITE,   size=14)
    data_font    = Font(name="Calibri",             color="FF334455", size=9)
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Border(
        left=Side(style="thin", color="FFD0D8E8"),
        right=Side(style="thin", color="FFD0D8E8"),
        bottom=Side(style="thin", color="FFD0D8E8"),
    )

    def style_header_row(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

    def style_data_row(ws, row, col_count, is_violation=False, vtype=""):
        color = LIGHT_ROW if row % 2 == 0 else WHITE
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=color)
            if c == 5 and is_violation:  # violation type column
                cell.font = Font(name="Calibri", bold=True,
                                  color=RED if "SPEED" in vtype else AMBER, size=9)
            else:
                cell.font = data_font
            cell.alignment = center
            cell.border = thin

    # ── Sheet 1: Violations ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Violations"

    # Title row
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "NTPC SMART SURVEILLANCE SYSTEM — VIOLATION REPORT"
    ws1["A1"].font   = title_font
    ws1["A1"].fill   = PatternFill("solid", fgColor=DARK_BLUE)
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 30

    # Subtitle
    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(violations)} violations"
    ws1["A2"].font = Font(name="Calibri", color="FF6688AA", italic=True, size=9)
    ws1["A2"].fill = PatternFill("solid", fgColor="FF0A1628")
    ws1["A2"].alignment = center

    # Headers
    headers = ["#", "Vehicle ID", "Vehicle Type", "License Plate", "Violation Type",
               "Speed (km/h)", "Speed Limit", "Date", "Time", "Camera ID"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=h)
    style_header_row(ws1, 3, len(headers))
    ws1.row_dimensions[3].height = 18

    # Data rows
    for i, v in enumerate(violations):
        row = i + 4
        vtype = v.get("violation_type", "")
        ws1.cell(row=row, column=1,  value=i + 1)
        ws1.cell(row=row, column=2,  value=v.get("vehicle_id", ""))
        ws1.cell(row=row, column=3,  value=v.get("vehicle_type", "").upper())
        ws1.cell(row=row, column=4,  value=v.get("plate_text", "UNKNOWN"))
        ws1.cell(row=row, column=5,  value=vtype)
        ws1.cell(row=row, column=6,  value=round(v["speed"], 1) if v.get("speed") else "N/A")
        ws1.cell(row=row, column=7,  value=v.get("speed_limit", "N/A"))
        ws1.cell(row=row, column=8,  value=v.get("date", ""))
        ws1.cell(row=row, column=9,  value=v.get("time", ""))
        ws1.cell(row=row, column=10, value=v.get("camera_id", ""))
        style_data_row(ws1, row, len(headers), is_violation=True, vtype=vtype)

    # Column widths
    for col, width in zip(range(1, len(headers)+1),
                          [6, 12, 14, 18, 16, 14, 12, 14, 10, 14]):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # Freeze top rows
    ws1.freeze_panes = "A4"

    # ── Sheet 2: All Vehicles ──────────────────────────────────────────────────
    if vehicles:
        ws2 = wb.create_sheet("All Vehicles")
        ws2.merge_cells("A1:H1")
        ws2["A1"] = "ALL DETECTED VEHICLES LOG"
        ws2["A1"].font   = title_font
        ws2["A1"].fill   = PatternFill("solid", fgColor=DARK_BLUE)
        ws2["A1"].alignment = center
        ws2.row_dimensions[1].height = 28

        headers2 = ["#", "Vehicle ID", "Type", "Plate", "Confidence", "Speed", "Date", "Time"]
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=2, column=col, value=h)
        style_header_row(ws2, 2, len(headers2))

        for i, v in enumerate(vehicles):
            row = i + 3
            ws2.cell(row=row, column=1, value=i + 1)
            ws2.cell(row=row, column=2, value=v.get("vehicle_id", ""))
            ws2.cell(row=row, column=3, value=v.get("vehicle_type", "").upper())
            ws2.cell(row=row, column=4, value=v.get("plate_text", ""))
            ws2.cell(row=row, column=5, value=f"{v.get('plate_conf',0):.0%}")
            ws2.cell(row=row, column=6, value=f"{v['speed']:.1f}" if v.get("speed") else "N/A")
            ws2.cell(row=row, column=7, value=v.get("date", ""))
            ws2.cell(row=row, column=8, value=v.get("time", ""))
            style_data_row(ws2, row, len(headers2))

        for col, width in zip(range(1, 9), [6, 12, 12, 18, 12, 12, 14, 10]):
            ws2.column_dimensions[get_column_letter(col)].width = width
        ws2.freeze_panes = "A3"

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "REPORT SUMMARY"
    ws3["A1"].font  = title_font
    ws3["A1"].fill  = PatternFill("solid", fgColor=DARK_BLUE)
    ws3["A1"].alignment = center
    ws3.row_dimensions[1].height = 28

    speeds = [v["speed"] for v in violations if v.get("speed")]
    summary = [
        ("Total Violations",      len(violations)),
        ("Overspeed Violations",  sum(1 for v in violations if v.get("violation_type") == "OVERSPEED")),
        ("No Helmet Violations",  sum(1 for v in violations if v.get("violation_type") == "NO_HELMET")),
        ("Total Vehicles Logged", len(vehicles)),
        ("Average Speed (km/h)",  round(sum(speeds)/len(speeds), 1) if speeds else 0),
        ("Max Speed (km/h)",      round(max(speeds), 1) if speeds else 0),
        ("Report Generated",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("System",                "NTPC Smart Surveillance System 2025"),
    ]
    for i, (label, value) in enumerate(summary):
        row = i + 2
        ws3.cell(row=row, column=1, value=label).font = Font(bold=True, color="FF334455")
        ws3.cell(row=row, column=2, value=value).font = Font(color="FF1a6abf")
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 30

    wb.save(output_path)
    print(f"[Excel] Report saved: {output_path}")
    return output_path
